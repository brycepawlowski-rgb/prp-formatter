from flask import Flask, request, send_file, render_template_string
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import io, os, zipfile, re

app = Flask(__name__)

HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>A&K PRP Formatter</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Segoe UI',Arial,sans-serif;background:#eef0f5;min-height:100vh;display:flex;flex-direction:column;align-items:center}
header{width:100%;background:#1F3864;color:#fff;padding:16px 36px;display:flex;align-items:center;gap:14px}
.logo{font-size:21px;font-weight:800}.logo span{color:#F4B942}
.htitle{font-size:13px;font-weight:600;opacity:.9}
.hright{margin-left:auto;font-size:11px;opacity:.5}
main{width:100%;max-width:660px;padding:36px 20px 60px;display:flex;flex-direction:column;gap:20px}
.card{background:#fff;border-radius:10px;border:1px solid #dde2ea;overflow:hidden;box-shadow:0 1px 4px rgba(0,0,0,.05)}
.chdr{background:#1F3864;color:#fff;padding:11px 20px;font-size:13px;font-weight:700;display:flex;align-items:center;gap:8px}
.cbody{padding:22px 20px}
#dz{border:2px dashed #c2cad8;border-radius:8px;padding:42px 24px;text-align:center;cursor:pointer;transition:all .18s;background:#f8f9fc;user-select:none}
#dz:hover,#dz.over{border-color:#1F3864;background:#edf1f9}
#dz .icon{font-size:36px;margin-bottom:10px}
#dz h3{font-size:15px;color:#2d3748;margin-bottom:5px}
#dz p{font-size:13px;color:#718096}
.lnk{color:#1F3864;font-weight:700;text-decoration:underline;cursor:pointer}
#fi{display:none}
#pill{display:none;align-items:center;gap:12px;margin-top:14px;background:#edf1f9;border-radius:8px;padding:11px 15px}
.fn{font-size:13px;font-weight:700;color:#1a2a4a}
.fs{font-size:11px;color:#718096;margin-top:1px}
.rm{margin-left:auto;background:none;border:none;cursor:pointer;color:#a0aec0;font-size:20px;line-height:1;padding:0 4px;border-radius:4px;transition:color .15s}
.rm:hover{color:#e53e3e}
#btn{width:100%;margin-top:18px;background:#1F3864;color:#fff;border:none;border-radius:8px;padding:13px;font-size:15px;font-weight:700;cursor:pointer;transition:background .18s;display:flex;align-items:center;justify-content:center;gap:9px}
#btn:hover:not(:disabled){background:#283f6e}
#btn:disabled{opacity:.4;cursor:not-allowed}
#pw{display:none;flex-direction:column;gap:7px;margin-top:16px}
.pt{background:#e2e8f0;border-radius:99px;height:7px;overflow:hidden}
.pf{background:#F4B942;height:100%;border-radius:99px;width:0%;transition:width .35s ease}
#pl{font-size:12px;color:#718096}
.st{display:none;border-radius:8px;padding:12px 16px;font-size:13px;margin-top:14px;align-items:center;gap:10px;font-weight:500}
.ok{background:#f0fff4;color:#276749;border:1px solid #c6f6d5}
.er{background:#fff5f5;color:#c53030;border:1px solid #fed7d7}
#dl{display:none;width:100%;margin-top:12px;background:#F4B942;color:#1a2040;border:none;border-radius:8px;padding:13px;font-size:15px;font-weight:800;cursor:pointer;transition:background .18s;align-items:center;justify-content:center;gap:9px}
#dl:hover{background:#e0a630}
.legend{display:grid;grid-template-columns:1fr 1fr;gap:12px}
.leg{display:flex;gap:10px;align-items:flex-start}
.sw{width:18px;height:18px;border-radius:3px;flex-shrink:0;margin-top:2px}
.leg h5{font-size:12px;font-weight:700;color:#1a2a4a;margin-bottom:2px}
.leg p{font-size:11px;color:#718096;line-height:1.5}
footer{font-size:11px;color:#b0b8c8;padding:16px;text-align:center}
</style>
</head>
<body>
<header>
  <div class="logo">A<span>&K</span> Finishing</div>
  <span style="opacity:.3;margin:0 8px">|</span>
  <div class="htitle">PRP Report Formatter</div>
  <div class="hright">Plex Export → Formatted Report</div>
</header>
<main>
  <div class="card">
    <div class="chdr">📄 Upload Your Plex PRP Export</div>
    <div class="cbody">
      <div id="dz" onclick="document.getElementById('fi').click()"
           ondragover="ev(event,'over')" ondragleave="ev(event,'leave')" ondrop="ev(event,'drop')">
        <div class="icon">📊</div>
        <h3>Drag &amp; drop your Plex export here</h3>
        <p>or <span class="lnk">click to browse</span></p>
        <p style="margin-top:8px;font-size:11px;color:#b0bac8">.xlsx files exported from the Plex PRP screen</p>
      </div>
      <input type="file" id="fi" accept=".xlsx" onchange="pick(this.files[0])">
      <div id="pill">
        <div style="font-size:22px">📗</div>
        <div><div class="fn" id="fn"></div><div class="fs" id="fs"></div></div>
        <button class="rm" onclick="clr()">×</button>
      </div>
      <button id="btn" onclick="run()" disabled>⚙️ &nbsp;Format Report</button>
      <div id="pw"><div class="pt"><div class="pf" id="pf"></div></div><div id="pl">Processing…</div></div>
      <div class="st" id="st"></div>
      <button id="dl">⬇️ &nbsp;Download Formatted Report</button>
    </div>
  </div>
  <div class="card">
    <div class="chdr">🎨 Color Guide</div>
    <div class="cbody">
      <div class="legend">
        <div class="leg"><div class="sw" style="background:#F4B942"></div><div><h5>Orange — Near-term positive</h5><p>Net is positive in near-term weeks. Stock on hand, demand approaching.</p></div></div>
        <div class="leg"><div class="sw" style="background:#F4CCCC;border:1px solid #cc9999"></div><div><h5>Red — Shortage</h5><p>Negative Net. Production required to avoid a stockout.</p></div></div>
        <div class="leg"><div class="sw" style="background:#1F3864"></div><div><h5>Navy — Headers</h5><p>Column headers matching the Plex PRP screen.</p></div></div>
        <div class="leg"><div class="sw" style="background:#EEF3FB;border:1px solid #b0c4de"></div><div><h5>Blue — Customer Releases</h5><p>Demand row per part for quick visual scanning.</p></div></div>
      </div>
    </div>
  </div>
</main>
<footer>A&amp;K Finishing LLC &middot; PRP Formatter &middot; Files are processed securely and never stored</footer>
<script>
let sf=null,dlUrl=null;
function ev(e,t){e.preventDefault();const z=document.getElementById('dz');if(t==='over')z.classList.add('over');if(t==='leave')z.classList.remove('over');if(t==='drop'){z.classList.remove('over');const f=e.dataTransfer.files[0];if(f&&f.name.endsWith('.xlsx'))pick(f);else show('Please drop an .xlsx file.',false);}}
function pick(f){sf=f;document.getElementById('fn').textContent=f.name;document.getElementById('fs').textContent=(f.size/1024).toFixed(1)+' KB';document.getElementById('pill').style.display='flex';document.getElementById('btn').disabled=false;document.getElementById('st').style.display='none';document.getElementById('dl').style.display='none';}
function clr(){sf=null;document.getElementById('fi').value='';document.getElementById('pill').style.display='none';document.getElementById('btn').disabled=true;document.getElementById('st').style.display='none';document.getElementById('dl').style.display='none';}
function show(msg,ok){const el=document.getElementById('st');el.textContent=msg;el.className='st '+(ok?'ok':'er');el.style.display='flex';}
function prog(p,l){document.getElementById('pf').style.width=p+'%';document.getElementById('pl').textContent=l;}
async function run(){
  if(!sf)return;
  document.getElementById('btn').disabled=true;
  document.getElementById('dl').style.display='none';
  document.getElementById('st').style.display='none';
  document.getElementById('pw').style.display='flex';
  prog(20,'Uploading…');
  const fd=new FormData();fd.append('file',sf);
  try{
    prog(50,'Formatting…');
    const r=await fetch('/format',{method:'POST',body:fd});
    if(!r.ok){throw new Error(await r.text());}
    prog(90,'Preparing download…');
    const blob=await r.blob();
    dlUrl=URL.createObjectURL(blob);
    prog(100,'Done!');
    setTimeout(()=>{
      document.getElementById('pw').style.display='none';
      show('✓ Report formatted successfully!',true);
      const dlBtn=document.getElementById('dl');
      dlBtn.style.display='flex';
      dlBtn.onclick=()=>{const a=document.createElement('a');a.href=dlUrl;a.download=sf.name.replace('.xlsx','')+'_formatted.xlsx';a.click();};
      document.getElementById('btn').disabled=false;
    },300);
  }catch(e){
    document.getElementById('pw').style.display='none';
    show('Error: '+e.message,false);
    document.getElementById('btn').disabled=false;
  }
}
</script>
</body>
</html>"""

def read_plex_excel(file_bytes):
    """Read Plex xlsx, patching the corrupt numFmtId='undefined' bug in Plex exports."""
    buf_in = io.BytesIO(file_bytes)
    buf_out = io.BytesIO()
    with zipfile.ZipFile(buf_in, 'r') as zin:
        with zipfile.ZipFile(buf_out, 'w', zipfile.ZIP_DEFLATED) as zout:
            for name in zin.namelist():
                data = zin.read(name)
                if name == 'xl/styles.xml':
                    text = data.decode('utf-8')
                    text = re.sub(r'numFmtId="undefined"', 'numFmtId="0"', text)
                    data = text.encode('utf-8')
                zout.writestr(name, data)
    buf_out.seek(0)
    return pd.read_excel(buf_out)

def format_prp(file_bytes):
    NAVY='1F3864';WHT='FFFFFF';LGRY='F2F2F2';ORG='F4B942'
    REDB='F4CCCC';REDF='CC0000';CRBG='EEF3FB';BDR='BFBFBF'
    def tf(h): return PatternFill('solid',fgColor=h,start_color=h)
    def tb():
        s=Side(style='thin',color=BDR)
        return Border(left=s,right=s,top=s,bottom=s)
    def si(v):
        try: return int(float(str(v).strip()))
        except: return 0

    df = read_plex_excel(file_bytes)
    df['Customer Part']=df['Customer Part'].astype(str).str.replace(r'_x000D_\s*','; ',regex=True)
    df['Related Parts']=df['Related Parts'].astype(str).str.replace(',',', ')

    cols=list(df.columns)
    si_end=cols.index('Non Useable Inventory')
    DATE_COLS=cols[si_end+2:]
    DORD=['Customer Releases','Net Demand','Scheduled','Net']
    NCOLS=13+len(DATE_COLS)
    today=datetime.today().strftime('%-m/%-d/%Y')

    wb=Workbook(); ws=wb.active; ws.title='PRP'

    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=NCOLS)
    tc=ws.cell(1,1,f'A&K Finishing LLC  |  PRODUCTION REQUIREMENTS PLANNING (PRP)  |  Workcenter: Line 1  |  Window: 8 Weeks  |  {today}')
    tc.font=Font(name='Arial',bold=True,size=10,color=WHT)
    tc.fill=tf(NAVY); tc.alignment=Alignment(horizontal='left',vertical='center')
    ws.row_dimensions[1].height=22

    HDRS=['Part No','Name','Std Job\nQty','Customer Part','Related Parts',
          'Lead\nTime','Min\nInv','Max\nInv','Rework\nInventory',
          'Inv\nWIP','Inv\nFG','Non Useable\nInventory','Demand Type']+list(DATE_COLS)
    ws.row_dimensions[2].height=40
    for ci,h in enumerate(HDRS,1):
        c=ws.cell(2,ci,h)
        c.font=Font(name='Arial',bold=True,color=WHT,size=9)
        c.fill=tf(NAVY); c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
        c.border=tb()

    WIDTHS=[16,26,8,22,30,7,8,8,9,8,8,12,18]+[10]*len(DATE_COLS)
    for i,w in enumerate(WIDTHS,1):
        ws.column_dimensions[get_column_letter(i)].width=w

    all_parts=df['Part No'].drop_duplicates().tolist()
    row=3
    for pi,pn in enumerate(all_parts):
        grp=df[df['Part No']==pn]; info=grp.iloc[0]
        dmap={r['Demand Type']:r for _,r in grp.iterrows()}
        fr=row; lr=row+3
        ibg='FFFFFF' if pi%2==0 else 'F7F9FC'

        for di,dtype in enumerate(DORD):
            cur=row+di
            ws.row_dimensions[cur].height=42 if di==0 else 15
            rbg=CRBG if dtype=='Customer Releases' else LGRY if dtype=='Net' else 'FFFFFF'

            def wc(col,val,bg,bold=False,italic=False,color='000000',wrap=False,align='left',vtop=False,_cur=cur):
                cell=ws.cell(_cur,col,val)
                cell.font=Font(name='Arial',size=9,bold=bold,italic=italic,color=color)
                cell.fill=tf(bg)
                cell.alignment=Alignment(horizontal=align,vertical='top' if vtop else 'center',wrap_text=wrap)
                cell.border=tb()

            if di==0:
                wc(1,str(info['Part No']),ibg,bold=True,vtop=True)
                wc(2,str(info['Name']),ibg,wrap=True,vtop=True)
                wc(3,si(info['Standard Job Quantity']),ibg,align='right')
                wc(4,str(info['Customer Part']),ibg,wrap=True,vtop=True)
                wc(5,str(info['Related Parts']),ibg,wrap=True,vtop=True)
                wc(6,si(info['Lead Time']),ibg,align='right')
                wc(7,si(info['Min Inv']),ibg,align='right')
                wc(8,si(info['Max Inv']),ibg,align='right')
                wc(9,si(info['Rework Inventory']),ibg,align='right')
                wc(10,si(info['Inv WIP']),ibg,align='right')
                wc(11,si(info['Inv FG']),ibg,align='right')
                wc(12,si(info['Non Useable Inventory']),ibg,align='right')
            else:
                for c in range(1,13):
                    cell=ws.cell(cur,c); cell.fill=tf(ibg); cell.border=tb()

            dt=ws.cell(cur,13,dtype)
            dt.font=Font(name='Arial',size=9,bold=(dtype=='Net'),italic=(dtype=='Net Demand'))
            dt.fill=tf(rbg); dt.alignment=Alignment(horizontal='left',vertical='center'); dt.border=tb()

            rd=dmap.get(dtype)
            for ci2,dcol in enumerate(DATE_COLS):
                col=14+ci2; val=si(rd[dcol]) if rd is not None else 0
                if dtype=='Net':
                    if val<0: bg,fg,bold=REDB,REDF,True
                    elif val>0 and col<=17: bg,fg,bold=ORG,'000000',True
                    else: bg,fg,bold=LGRY,'000000',True
                else: bg,fg,bold=rbg,'000000',False
                nc=ws.cell(cur,col,val)
                nc.font=Font(name='Arial',size=9,bold=bold,color=fg)
                nc.fill=tf(bg); nc.alignment=Alignment(horizontal='right',vertical='center')
                nc.border=tb(); nc.number_format='#,##0;-#,##0;0'

        for col in range(1,13):
            ws.merge_cells(start_row=fr,start_column=col,end_row=lr,end_column=col)
            ws.cell(fr,col).border=tb()
        row+=4

    ws.freeze_panes='N3'
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

@app.route('/')
def index(): return render_template_string(HTML)

@app.route('/format', methods=['POST'])
def fmt():
    f=request.files.get('file')
    if not f: return 'No file uploaded',400
    try:
        buf=format_prp(f.read())
        return send_file(buf,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True,download_name='PRP_formatted.xlsx')
    except Exception as e: return str(e),500

if __name__=='__main__':
    port=int(os.environ.get('PORT',5050))
    app.run(host='0.0.0.0',port=port)
